#!/usr/bin/env python
"""Parse the DLGS Property Tax Tables xlsx and rewrite constants.py.

Populates TAX_RATE_PER_HUNDRED, TOTAL_LEVY, and the six-key LEVY_BREAKDOWN
(municipal, county, library, local_school, regional_school, open_space).
"""
from __future__ import annotations

import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path


COMPONENT_HEADER_PATTERNS: dict[str, list[str]] = {
    "municipal": ["municipal", "muni"],
    "county": ["county"],
    "library": ["library"],
    "local_school": ["local school", "local district school"],
    "regional_school": ["regional school"],
    "open_space": ["open space"],
    "tax_rate": ["general tax rate", "total tax rate", "rate per 100", "rate per $100"],
    "total_levy": ["total levy", "total general tax levy", "amount to be raised"],
}

CONSTANTS_PATH = Path("src/fairhaven_tax/constants.py")


def _to_decimal(v) -> Decimal | None:
    if v is None:
        return None
    try:
        s = str(v).strip().replace("$", "").replace(",", "")
        if s.lower() in {"nan", "none", ""}:
            return None
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _matches(cell, patterns: list[str]) -> bool:
    if cell is None:
        return False
    s = str(cell).lower().strip()
    return any(p in s for p in patterns)


def _latest_snapshot() -> Path:
    base = Path("data/raw/dlgs_tax_tables")
    if not base.exists():
        raise FileNotFoundError(
            f"missing snapshot dir: {base}. Run `make acquire-dlgs` first."
        )
    snaps = sorted([d for d in base.iterdir() if d.is_dir()])
    if not snaps:
        raise FileNotFoundError(f"no snapshots in {base}")
    return snaps[-1]


def _find_workbook(snap: Path) -> Path:
    for ext in ("*.xlsx", "*.xls"):
        for p in snap.glob(ext):
            return p
    raise FileNotFoundError(f"no .xls/.xlsx in {snap}")


def _scan_workbook(path: Path) -> dict[str, Decimal]:
    """Return dict with keys: tax_rate, total_levy, municipal, county, library,
    local_school, regional_school, open_space (all Decimal)."""
    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, max_row=20, values_only=True))
        for hdr_idx, hdr in enumerate(rows):
            if hdr is None:
                continue
            # find a row that looks like the column header (contains "municipal" or "muni")
            header_has_muni = any(
                _matches(c, COMPONENT_HEADER_PATTERNS["municipal"]) for c in hdr
            )
            if not header_has_muni:
                continue
            # Capture column indices
            col_idx: dict[str, int] = {}
            for i, c in enumerate(hdr):
                for key, pats in COMPONENT_HEADER_PATTERNS.items():
                    if key not in col_idx and _matches(c, pats):
                        col_idx[key] = i
            if "municipal" not in col_idx:
                continue
            # Now scan the rest of the sheet for a row containing "fair haven"
            for data_row in ws.iter_rows(min_row=hdr_idx + 2, values_only=True):
                if data_row is None:
                    continue
                row_text = " ".join(str(c) for c in data_row if c is not None).lower()
                if "fair haven" not in row_text:
                    continue
                values: dict[str, Decimal | None] = {}
                for key, idx in col_idx.items():
                    if idx >= len(data_row):
                        values[key] = None
                        continue
                    values[key] = _to_decimal(data_row[idx])
                # Validate required minima
                if values.get("tax_rate") is None:
                    continue  # try next match row
                return {k: v for k, v in values.items() if v is not None}
    raise ValueError(
        f"could not locate Fair Haven row in {path}. "
        "If DLGS column headers drifted, update COMPONENT_HEADER_PATTERNS in scripts/extract_dlgs.py."
    )


def _rewrite_constants(values: dict[str, Decimal]) -> None:
    if "tax_rate" not in values:
        raise ValueError("tax_rate missing from extracted values")
    text = CONSTANTS_PATH.read_text()
    tax_rate = values["tax_rate"]
    # Sanity check
    if not (Decimal("1.2") <= tax_rate <= Decimal("2.0")):
        print(f"WARN: tax_rate={tax_rate} outside expected $1.20-$2.00 range")

    text = re.sub(
        r"^TAX_RATE_PER_HUNDRED:.*$",
        f'TAX_RATE_PER_HUNDRED: Decimal | None = Decimal("{tax_rate}")',
        text,
        flags=re.MULTILINE,
    )
    if "total_levy" in values:
        text = re.sub(
            r"^TOTAL_LEVY:.*$",
            f'TOTAL_LEVY: Decimal | None = Decimal("{values["total_levy"]}")',
            text,
            flags=re.MULTILINE,
        )
    breakdown_keys = ["municipal", "county", "library", "local_school", "regional_school", "open_space"]
    if any(k in values for k in breakdown_keys):
        parts = []
        for k in breakdown_keys:
            if k in values:
                parts.append(f'"{k}": Decimal("{values[k]}")')
        breakdown_repr = "{" + ", ".join(parts) + "}"
        text = re.sub(
            r"^LEVY_BREAKDOWN:.*$",
            f"LEVY_BREAKDOWN: dict[str, Decimal] | None = {breakdown_repr}",
            text,
            flags=re.MULTILINE,
        )
    CONSTANTS_PATH.write_text(text)


def main() -> int:
    try:
        snap = _latest_snapshot()
        wb_path = _find_workbook(snap)
        from fairhaven_tax.ingest.manifest import verify_manifest
        ok, errors = verify_manifest(snap)
        if not ok:
            print(f"ERROR: manifest verification failed in {snap}:", file=sys.stderr)
            for err in errors:
                print(f"  - {err}", file=sys.stderr)
            return 2
        values = _scan_workbook(wb_path)
        _rewrite_constants(values)
        print("Extracted DLGS values:")
        for k, v in values.items():
            print(f"  {k}: {v}")
        return 0
    except (FileNotFoundError, ValueError) as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    sys.exit(main())
